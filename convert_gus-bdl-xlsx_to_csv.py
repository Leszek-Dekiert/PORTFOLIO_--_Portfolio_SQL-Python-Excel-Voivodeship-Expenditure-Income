# -*- coding: utf-8 -*-

"""

Requirements:
1. Python 3.9.7 or higher
2. openpyxl library
3. pathlib library

Source of data: GUS BDL (https://bdl.stat.gov.pl/)

Author: Leszek Dekiert
Formatted with: Black 25.1.0

Code description below


I. Extraction and conversion from xlsx to CSV_part_1

Preperation of xlsx files in Excel: the source data was cleaned and 
converted so as to leave only the features that really differentiate in the 
column headers: e.g. when selecting the entire scope of specific thematic 
data on the GUS BDL website, the headers remain in the xlsx file as 
row 1: "general", row 2: "general", where in another case 
(selecting more specific data) it would be e.g. row 1: "women", 
row 2: "unemployed". In addition, all the necessary and cleaned data were 
placed in one workbook, divided into sheets.

Code will extract and convert certain data from xlsx file into CSV file.
This data is a GUS BDL type of data extracted from website:
https://bdl.stat.gov.pl/bdl/dane/podgrup/temat.
The type of GUS BDL data that can be converted in this case has been prepared
to have two first columns: code (A) and region (B).
The first row starting from C are years.
User of this code needs to input: data file path, worksheet of the file,
name of new CSV file.
After that process is concluded its recomended to use import feature of Excel
to put all CSV files together for further SQL analisys.
For a different type of data please use the other code stated below.
This code will work with sheets 1, 2, 3, 4 of file: 
"voivodeship_expenditure_income_raw-data-for-conversion"

"""

# Importing load_workbook from library openpyxl and Path from pathlib
from openpyxl import load_workbook
from pathlib import Path

# Choosing file for conversion
data_file = input(
    "Please enter the path to the file You would like to "
    'convert. E.g.: "C:/Users/Leszek/extracting_from_excel/data/'
    'Budzet_wojewodztw.xlsx": '
)

# Load the entire workbook
workbook = load_workbook(data_file)

# List all the sheets in the file. Appending sheetnames on a list
worksheets = []
print("Found the following worksheets:")
for sheetname in workbook.sheetnames:
    print(sheetname)
    worksheets.append(sheetname)

# Load one column and append to a list. Choosing worksheet for convertion
worksheet_str = input(
    "Please enter a number of worksheet You would like "
    "to convert. Remember to use a number that corresponds "
    "with order of Your worksheet in the file: "
)
worksheet_num = int(worksheet_str) - 1
sheet = workbook[worksheets[int(worksheet_num)]]
all_rows = list(sheet.rows)
row_num = 0

# Prepare data for later iterations. Removing two useless column names:
#'Kod' and 'Nazwa'
records_code = []

for row in sheet:
    record = row[row_num].value
    if record == "Kod":
        continue
    else:
        records_code.append(record)

print(records_code)

records_name = []

for row in sheet:
    record = row[row_num + 1].value
    if record == "Nazwa":
        continue
    else:
        records_name.append(record)

print(records_name)

records_year = []

for cell in all_rows[0]:
    if cell.value == "Nazwa" or cell.value == "Kod":
        continue
    else:
        records_year.append(cell.value)

print(records_year)
# j = purpose of j is to not append first object
records_value = []
row_num = 2
j = 0

try:
    while row_num < 9999:
        for row in sheet:
            record = row[row_num].value
            j += 1
            if j == 1:
                continue
            else:
                records_value.append(record)
                print(record)
        row_num += 1
        j = 0
    print(records_value)
except IndexError:
    pass

# Getting specific cells and appending to CSV file. Naming new CSV file.
CSV_file_name = input("Enter the name of new CSV file that You are creating: ")
CSV_file_path = Path(data_file)
CSV_folder_path = CSV_file_path.parent
full_path = f"{CSV_folder_path}/{CSV_file_name}"

try:
    with open(f"{full_path}.txt", "a") as file:
        record_code = 0
        record_name = 0
        record_year = 0
        record_value = 0
        number_of_regions = 17
        while record_year != 999:
            while record_value != number_of_regions:
                full_record = (
                    str(records_code[record_code])
                    + ","
                    + str(records_name[record_name])
                    + ","
                    + str(records_year[record_year])
                    + ","
                    + str(records_value[record_value])
                )
                file.write(full_record + "\n")
                record_value += 1
                record_code += 1
                record_name += 1
            record_code = 0
            record_name = 0
            record_year += 1
            number_of_regions += 17
except IndexError:
    pass
# %%

"""


II. Extraction and conversion from xlsx to CSV_part_2

Preperation of xlsx files in Excel: the source data was cleaned and 
converted so as to leave only the features that really differentiate in the 
column headers: e.g. when selecting the entire scope of specific thematic 
data on the GUS BDL website, the headers remain in the xlsx file as 
row 1: "general", row 2: "general", where in another case 
(selecting more specific data) it would be e.g. row 1: "women", 
row 2: "unemployed". In addition, all the necessary and cleaned data were 
placed in one workbook, divided into sheets.

Code will extract and convert certain data from xlsx file into CSV file.
This data is a GUS BDL type of data extracted from website:
https://bdl.stat.gov.pl/bdl/dane/podgrup/temat.
The type of GUS BDL data that can be converted in this case has been prepared
to have two first columns: code (A) and region (B).
The first row starting from C1 are additional traits.
The second row starting from C2 are years.
User of this code needs to input: data file path, worksheet of the file,
name of new CSV file.
After that process is concluded its recomended to use import feature of Excel
to put all CSV files together for further SQL analisys.
For a different type of data please use the other code stated above.,
This code will work with sheets 5, 6 of file: 
"voivodeship_expenditure_income_raw-data-for-conversion"

"""

# Importing load_workbook from library openpyxl
from openpyxl import load_workbook

# Choosing file for conversion
data_file = input(
    "Please enter the path to the file You would like to "
    'convert. E.g.: "C:/Users/Leszek/extracting_from_excel/data/'
    'Budzet_wojewodztw.xlsx": '
)

# Load the entire workbook.
workbook = load_workbook(data_file)

# List all the sheets in the file. Appending sheetnames on a list
worksheets = []
print("Found the following worksheets:")
for sheetname in workbook.sheetnames:
    print(sheetname)
    worksheets.append(sheetname)

# Load one column and append to a list. Choosing worksheet for convertion
worksheet_str = input(
    "Please enter a number of worksheet You would like "
    "to convert. Remember to use a number that corresponds "
    "with order of Your worksheet in the file: "
)
worksheet_num = int(worksheet_str) - 1
sheet = workbook[worksheets[int(worksheet_num)]]
all_rows = list(sheet.rows)
row_num = 0

# Prepare data for later iterations. Removing two useless column names:
#'Kod', 'Nazwa' and 'ND'
records_code = []

for row in sheet:
    record = row[row_num].value
    if record == "Kod" or record == "ND":
        continue
    else:
        records_code.append(record)

print(records_code)

records_name = []

for row in sheet:
    record = row[row_num + 1].value
    if record == "Nazwa" or record == "ND":
        continue
    else:
        records_name.append(record)

print(records_name)

records_year = []

for cell in all_rows[1]:
    if cell.value == "Nazwa" or cell.value == "Kod" or cell.value == "ND":
        continue
    else:
        records_year.append(cell.value)

print(records_year)

# j = purpose of j is to not append first two objects
records_value = []
row_num = 2
j = 0

try:
    while row_num < 9999:
        for row in sheet:
            record = row[row_num].value
            j += 1
            if j == 1 or j == 2:
                continue
            else:
                records_value.append(record)
                print(record)
        row_num += 1
        j = 0
    print(records_value)
except IndexError:
    pass

records_trait = []

for cell in all_rows[0]:
    if (
        cell.value == "Nazwa"
        or cell.value == "Kod"
        or cell.value == "ND"
        or cell.value.startswith("=")
    ):
        continue
    else:
        records_trait.append(cell.value)

print(records_trait)

# Getting specific cells and appending to CSV file.
CSV_file_name = input("Enter the name of new CSV file that You are creating: ")
CSV_file_path = Path(data_file)
CSV_folder_path = CSV_file_path.parent
full_path = f"{CSV_folder_path}/{CSV_file_name}"

try:
    with open(f"{full_path}.txt", "a") as file:
        record_code = 0
        record_name = 0
        record_year = 0
        record_value = 0
        rt = 0
        number_of_regions = 17
        trait_number = 0
        while record_year != 9999:
            while record_value != number_of_regions:
                full_record = (
                    str(records_code[record_code])
                    + ","
                    + str(records_name[record_name])
                    + ","
                    + str(records_year[record_year])
                    + ","
                    + str(records_value[record_value])
                    + ","
                    + str(records_trait[rt])
                )
                file.write(full_record + "\n")
                record_value += 1
                record_code += 1
                record_name += 1
            record_code = 0
            record_name = 0
            record_year += 1
            number_of_regions += 17
            trait_number += 1
            if trait_number == 16:
                rt += 1
                trait_number = 0
except IndexError:
    pass
