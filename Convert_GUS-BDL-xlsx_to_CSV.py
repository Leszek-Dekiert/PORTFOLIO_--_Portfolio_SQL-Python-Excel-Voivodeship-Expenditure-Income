# -*- coding: utf-8 -*-


###

#Source of data: GUS BDL (https://bdl.stat.gov.pl/)
#Code description below

###

#Extraction and conversion from xlsx to CSV_part_1

###


#Code will extract and convert certain data from xlsx file into CSV file.
#This data is a GUS BDL type of data extracted from website:
#https://bdl.stat.gov.pl/bdl/dane/podgrup/temat.
#The type of GUS BDL data that can be converted in this case has been prepared
#to have two first columns: code (A) and region (B). 
#The first row starting from C are years.
#User of this code needs to input: data file path, worksheet of the file,
#name of new CSV file.
#After that process is concluded its recomended to use import feature of Excel
#to put all CSV files together for further SQL analisys.
#For a different type of data please use the other code stated below.

###


#Importing load_workbook from library openpyxl
from openpyxl import load_workbook

#Choosing file for conversion
data_file = input('Please enter the path to the file You would like to '
                  'convert. E.g.: "C:/Users/Leszek/extracting_from_excel/data/'
                  'Budzet_wojewodztw.xlsx": ' )

#Load the entire workbook
workbook = load_workbook(data_file)

#List all the sheets in the file. Appending sheetnames on a list
worksheets = []
print("Found the following worksheets:")
for sheetname in workbook.sheetnames:
    print(sheetname)
    worksheets.append(sheetname)

#Load one column and append to a list. Choosing worksheet for convertion
worksheet_str = input('Please enter a number of worksheet You would like '
                      'to convert. Remember to use a number that corresponds '
                      'with order of Your worksheet in the file: ')
worksheet_num = int(worksheet_str) - 1
sheet = workbook[worksheets[int(worksheet_num)]]
all_rows = list(sheet.rows)
row_num = 0

#Prepare data for later iterations. Removing two useless column names:
#'Kod' and 'Nazwa'
records_code = []

for row in sheet:
    record = row[row_num].value
    if record == 'Kod':
        continue
    else:
        records_code.append(record)
  
print(records_code) 

records_name = []

for row in sheet:
    record = row[row_num + 1].value
    if record == 'Nazwa':
        continue
    else:
        records_name.append(record)
  
print(records_name)

records_year = []

for cell in all_rows[0]:
    if cell.value == 'Nazwa' or cell.value == 'Kod':
        continue
    else:
        records_year.append(cell.value)
  
print(records_year)
#j = purpose of j is to not append first object
records_value = []
row_num = 2
j = 0

try:
    while row_num < 9999:
        for row in sheet:
                record = row[row_num].value
                j += 1
                if j == 1:
                    continue
                else:
                    records_value.append(record)
                    print(record)
        row_num += 1
        j = 0
    print(records_value)
except IndexError:
    pass

#Getting specific cells and appending to CSV file. Naming new CSV file.
#rc = records code,
#rn = records name,
#ry = records year,
#rv = record value,
#s = number of regions corresponding with values after which you move to
#next column
CSV_file_name = input('Enter the name of new CSV file that You are creating: ')
try:
    with open(f'{CSV_file_name}.txt', 'a') as file:
            rc = 0
            rn = 0
            ry = 0
            rv = 0
            s = 17
            while ry != 999:
                while rv != s:
                    full_record = (str(records_code[rc]) + ',' + 
                                    str(records_name[rn]) + ',' + 
                                    str(records_year[ry]) + ',' + 
                                    str(records_value[rv]))
                    file.write(full_record + '\n')
                    rv += 1
                    rc += 1
                    rn += 1
                rc = 0
                rn = 0
                ry += 1
                s += 17
except IndexError:
    pass
# %%


###

#Source of data: GUS BDL (https://bdl.stat.gov.pl/)
#Code description below

###

#Extraction and conversion from xlsx to CSV_part_2

###


#Code will extract and convert certain data from xlsx file into CSV file.
#This data is a GUS BDL type of data extracted from website:
#https://bdl.stat.gov.pl/bdl/dane/podgrup/temat.
#The type of GUS BDL data that can be converted in this case has been prepared
#to have two first columns: code (A) and region (B). 
#The first row starting from C1 are additional traits.
#The second row starting from C2 are years.
#User of this code needs to input: data file path, worksheet of the file,
#name of new CSV file.
#After that process is concluded its recomended to use import feature of Excel
#to put all CSV files together for further SQL analisys.
#For a different type of data please use the other code stated below.

###


#Importing load_workbook from library openpyxl
from openpyxl import load_workbook

#Choosing file for conversion
data_file = input('Please enter the path to the file You would like to '
                  'convert. E.g.: "C:/Users/Leszek/extracting_from_excel/data/'
                  'Budzet_wojewodztw.xlsx": ' )

#Load the entire workbook.
workbook = load_workbook(data_file)

#List all the sheets in the file. Appending sheetnames on a list
worksheets = []
print("Found the following worksheets:")
for sheetname in workbook.sheetnames:
    print(sheetname)
    worksheets.append(sheetname)

#Load one column and append to a list. Choosing worksheet for convertion
worksheet_str = input('Please enter a number of worksheet You would like '
                      'to convert. Remember to use a number that corresponds '
                      'with order of Your worksheet in the file: ')
worksheet_num = int(worksheet_str) - 1
sheet = workbook[worksheets[int(worksheet_num)]]
all_rows = list(sheet.rows)
row_num = 0

#Prepare data for later iterations. Removing two useless column names:
#'Kod', 'Nazwa' and 'ND'
records_code = []

for row in sheet:
    record = row[row_num].value
    if record == 'Kod' or record == 'ND':
        continue
    else:
        records_code.append(record)
  
print(records_code) 

records_name = []

for row in sheet:
    record = row[row_num + 1].value
    if record == 'Nazwa' or record == 'ND':
        continue
    else:
        records_name.append(record)
  
print(records_name)

records_year = []

for cell in all_rows[1]:
    if cell.value == 'Nazwa' or cell.value == 'Kod' or cell.value == 'ND':
        continue
    else:
        records_year.append(cell.value)
  
print(records_year)

#j = purpose of j is to not append first two objects
records_value = []
row_num = 2
j = 0

try:
    while row_num < 9999:
        for row in sheet:
                record = row[row_num].value
                j += 1
                if j == 1 or j == 2:
                    continue
                else:
                    records_value.append(record)
                    print(record)
        row_num += 1
        j = 0
    print(records_value)
except IndexError:
    pass

records_trait = []

for cell in all_rows[0]:
    if (cell.value == 'Nazwa' or 
        cell.value == 'Kod' or 
        cell.value == 'ND' or 
        cell.value.startswith('=')):
        continue
    else:
            records_trait.append(cell.value)
  
print(records_trait)

#Getting specific cells and appending to CSV file. 
#rc = records code,
#rn = records name, 
#ry = records year, 
#rv = record value,
#s = number of regions corresponding with values after which you move to
#next column
#i = keeps track of traits
CSV_file_name = input('Enter the name of new CSV file that You are creating: ')
try:
    with open(f'{CSV_file_name}.txt', 'a') as file:
        rc = 0
        rn = 0
        ry = 0
        rv = 0
        rt = 0
        s = 17
        i = 0
        while ry != 9999:
            while rv != s:
                full_record = (str(records_code[rc]) + ',' + 
                                str(records_name[rn]) + ',' + 
                                str(records_year[ry]) + ',' + 
                                str(records_value[rv]) + ',' + 
                                str(records_trait[rt]))
                file.write(full_record + '\n')
                rv += 1
                rc += 1
                rn += 1
            rc = 0
            rn = 0
            ry += 1
            s += 17
            i += 1
            if i == 16:
                rt += 1
                i = 0
except IndexError:
    pass